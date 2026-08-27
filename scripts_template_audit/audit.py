#!/usr/bin/env python3
"""模板预检脚本:列出 Excel 模板中所有 Rust umya-spreadsheet 需要处理的元素。"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook


def cell_value_short(c):
    v = c.value
    if v is None:
        return None
    s = str(v)
    return s if len(s) <= 80 else s[:77] + "..."


def audit(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, keep_vba=False)
    report = {
        "file": str(path),
        "size_bytes": path.stat().st_size,
        "sheets": [],
        "defined_names": [],
    }

    # Defined names
    try:
        for n in wb.defined_names:
            dests = list(wb.defined_names[n].destinations) if n in wb.defined_names else []
            report["defined_names"].append({"name": n, "destinations": dests})
    except Exception as e:
        report["defined_names_error"] = str(e)

    for ws in wb.worksheets:
        sheet = {
            "name": ws.title,
            "state": ws.sheet_state,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_cells": [str(r) for r in ws.merged_cells.ranges],
            "formulas": [],
            "conditional_formatting": [],
            "data_validations": [],
            "print_area": ws.print_area,
            "print_titles": ws.print_titles,
            "page_setup": {
                "orientation": ws.page_setup.orientation,
                "paperSize": ws.page_setup.paperSize,
                "fitToWidth": ws.page_setup.fitToWidth,
                "fitToHeight": ws.page_setup.fitToHeight,
            },
            "page_margins": {
                "left": ws.page_margins.left,
                "right": ws.page_margins.right,
                "top": ws.page_margins.top,
                "bottom": ws.page_margins.bottom,
            },
            "header_footer": {
                "oddHeader": ws.oddHeader.center.text if ws.oddHeader and ws.oddHeader.center else None,
                "oddFooter": ws.oddFooter.center.text if ws.oddFooter and ws.oddFooter.center else None,
            },
            "charts_count": len(ws._charts) if hasattr(ws, "_charts") else 0,
            "images_count": len(ws._images) if hasattr(ws, "_images") else 0,
            "tables": [t for t in (ws.tables.keys() if hasattr(ws, "tables") else [])],
            "freeze_panes": ws.freeze_panes,
            "sheet_view_zoom": ws.sheet_view.zoomScale,
            "row_heights_sample": {},
            "col_widths_sample": {},
            "non_empty_cells_sample": [],
            "number_formats_used": set(),
            "font_names_used": set(),
            "fill_colors_used": set(),
        }

        # 公式扫描(全表,会比较大)
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and isinstance(c.value, str) and c.value.startswith("="):
                    sheet["formulas"].append({
                        "cell": c.coordinate,
                        "formula": c.value,
                    })

        # 条件格式
        try:
            for rng, rules in ws.conditional_formatting._cf_rules.items():
                sheet["conditional_formatting"].append({
                    "range": str(rng),
                    "rule_count": len(rules),
                    "rule_types": [r.type for r in rules],
                })
        except Exception as e:
            sheet["conditional_formatting_error"] = str(e)

        # 数据验证
        try:
            for dv in ws.data_validations.dataValidation:
                sheet["data_validations"].append({
                    "type": dv.type,
                    "formula1": dv.formula1,
                    "ranges": str(dv.sqref),
                })
        except Exception as e:
            sheet["data_validations_error"] = str(e)

        # 行列宽(只取显式设过的)
        for k, dim in ws.row_dimensions.items():
            if dim.height is not None:
                sheet["row_heights_sample"][str(k)] = dim.height
        for k, dim in ws.column_dimensions.items():
            if dim.width is not None:
                sheet["col_widths_sample"][str(k)] = dim.width

        # 非空单元格样本(扫描,前 60 个)
        cnt = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and cnt < 60:
                    entry = {
                        "cell": c.coordinate,
                        "value": cell_value_short(c),
                        "is_formula": isinstance(c.value, str) and c.value.startswith("="),
                        "number_format": c.number_format,
                    }
                    sheet["non_empty_cells_sample"].append(entry)
                    sheet["number_formats_used"].add(c.number_format)
                    if c.font and c.font.name:
                        sheet["font_names_used"].add(c.font.name)
                    if c.fill and c.fill.fgColor and c.fill.fgColor.rgb and c.fill.fgColor.rgb != "00000000":
                        sheet["fill_colors_used"].add(str(c.fill.fgColor.rgb))
                    cnt += 1

        # 转 set 为 sorted list
        sheet["number_formats_used"] = sorted(sheet["number_formats_used"])
        sheet["font_names_used"] = sorted(sheet["font_names_used"])
        sheet["fill_colors_used"] = sorted(sheet["fill_colors_used"])

        report["sheets"].append(sheet)

    return report


if __name__ == "__main__":
    for p in sys.argv[1:]:
        r = audit(Path(p))
        print(json.dumps(r, ensure_ascii=False, indent=2, default=str))
        print("\n" + "=" * 80 + "\n")
